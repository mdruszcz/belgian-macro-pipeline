"""
Belgian Macroeconomic Database
==============================
Fetches GDP and confidence data from NBB SDMX and DBnomics,
stores in SQLite, and exports to CSV/JSON.

Runs daily via GitHub Actions.
"""

import argparse
import csv
import io
import logging
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
import requests
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent / "scripts"))

from rename_legacy_tables import rename_legacy_tables  # noqa: E402

from src.db.migrate import run as run_migrations  # noqa: E402
from src.validation.config_schema import load_and_validate_all  # noqa: E402

# ─── Configuration ────────────────────────────────────────────────

DB_PATH = Path(__file__).parent / "data" / "belgian_macro.db"

NBB_CSV_HEADER = {"Accept": "application/vnd.sdmx.data+csv;version=2.0.0"}

FPB_XLSX_URL = "https://www.plan.be/sites/default/files/documents/FOR_BE_FR.xlsx"

CONFIG_DIR = Path(__file__).parent / "config"


def _load_sources() -> dict:
    """Rebuild the SOURCES-dict shape fetch_all()/upsert_indicator() expect,
    from config/indicators/*.yaml + config/sources/*.yaml, per
    docs/features/indicator_config.md. Indicators whose source's adapter is
    not "nbb"/"dbnomics" (i.e. adapter "fpb", the forecast pseudo-indicators)
    are excluded -- forecasts are fetched by FPBFetcher directly, exactly as
    before, and never belonged in this dict."""
    indicators, sources = load_and_validate_all(CONFIG_DIR / "indicators", CONFIG_DIR / "sources")
    out = {}
    for code, ind in indicators.items():
        source = sources[ind["source_id"]]
        if source["adapter"] not in ("nbb", "dbnomics"):
            continue
        out[code] = {
            "name": ind["name"]["en"],
            "url": source["base_url"] + ind["fetch"]["query"],
            "frequency": ind["frequency"],
            "unit": ind["unit"],
            "source_agency": source["agency"],
            "description": ind.get("description", {}).get("en", ""),
            "type": source["adapter"],
        }
    return out


SOURCES = _load_sources()

REQUEST_TIMEOUT = 30

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-7s | %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("belgian_macro")


# ─── Database ─────────────────────────────────────────────────────


class MacroDatabase:
    def __init__(self, db_path: Path = DB_PATH):
        self.db_path = db_path
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self.conn = sqlite3.connect(str(db_path))
        self.conn.execute("PRAGMA journal_mode=WAL")
        self._init_schema()

    def _init_schema(self):
        # Order matters: rename any legacy-shaped tables out of the way first,
        # then apply canonical migrations, then create the (now non-colliding,
        # renamed) legacy tables if they don't exist yet. See
        # scripts/rename_legacy_tables.py and docs/decisions/0001-data-model.md.
        rename_legacy_tables(self.conn)
        run_migrations(self.db_path)
        self.conn.executescript("""
            CREATE TABLE IF NOT EXISTS legacy_indicators (
                code          TEXT PRIMARY KEY,
                name          TEXT NOT NULL,
                frequency     TEXT NOT NULL,
                unit          TEXT NOT NULL,
                source_agency TEXT NOT NULL,
                description   TEXT,
                api_url       TEXT
            );
            CREATE TABLE IF NOT EXISTS legacy_observations (
                indicator_code TEXT NOT NULL,
                period         TEXT NOT NULL,
                value          REAL NOT NULL,
                obs_status     TEXT,
                fetched_at     TEXT NOT NULL,
                PRIMARY KEY (indicator_code, period),
                FOREIGN KEY (indicator_code) REFERENCES legacy_indicators(code)
            );
            CREATE TABLE IF NOT EXISTS legacy_fetch_log (
                id             INTEGER PRIMARY KEY AUTOINCREMENT,
                indicator_code TEXT NOT NULL,
                fetched_at     TEXT NOT NULL,
                rows_upserted  INTEGER NOT NULL,
                status         TEXT NOT NULL,
                message        TEXT
            );
            CREATE INDEX IF NOT EXISTS idx_obs_period
                ON legacy_observations(indicator_code, period DESC);
            CREATE TABLE IF NOT EXISTS forecasts (
                institution    TEXT NOT NULL,
                indicator      TEXT NOT NULL,
                year           TEXT NOT NULL,
                value          REAL,
                updated_at     TEXT,
                fetched_at     TEXT NOT NULL,
                PRIMARY KEY (institution, indicator, year)
            );
        """)
        self.conn.commit()

    def upsert_indicator(self, code: str, meta: dict):
        self.conn.execute(
            """
            INSERT INTO legacy_indicators (code, name, frequency, unit, source_agency, description, api_url)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(code) DO UPDATE SET
                name=excluded.name, frequency=excluded.frequency,
                unit=excluded.unit, source_agency=excluded.source_agency,
                description=excluded.description, api_url=excluded.api_url
        """,
            (
                code,
                meta["name"],
                meta["frequency"],
                meta["unit"],
                meta["source_agency"],
                meta.get("description", ""),
                meta.get("url", ""),
            ),
        )
        self.conn.commit()

    def upsert_observations(self, indicator_code: str, rows: list[dict]) -> int:
        now = datetime.now(timezone.utc).isoformat()
        for row in rows:
            self.conn.execute(
                """
                INSERT INTO legacy_observations (indicator_code, period, value, obs_status, fetched_at)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(indicator_code, period) DO UPDATE SET
                    value=excluded.value, obs_status=excluded.obs_status,
                    fetched_at=excluded.fetched_at
            """,
                (indicator_code, row["period"], row["value"], row.get("obs_status", ""), now),
            )
        self.conn.commit()
        return len(rows)

    def log_fetch(self, code: str, count: int, status: str, msg: str = ""):
        now = datetime.now(timezone.utc).isoformat()
        self.conn.execute(
            "INSERT INTO legacy_fetch_log (indicator_code, fetched_at, rows_upserted, status, message) VALUES (?,?,?,?,?)",
            (code, now, count, status, msg),
        )
        self.conn.commit()

    def get_latest(self, code: str) -> dict | None:
        cur = self.conn.execute(
            """
            SELECT o.period, o.value, o.obs_status, o.fetched_at, i.name, i.unit
            FROM legacy_observations o JOIN legacy_indicators i ON o.indicator_code = i.code
            WHERE o.indicator_code = ? ORDER BY o.period DESC LIMIT 1
        """,
            (code,),
        )
        r = cur.fetchone()
        if not r:
            return None
        return {
            "indicator_code": code,
            "period": r[0],
            "value": r[1],
            "obs_status": r[2],
            "fetched_at": r[3],
            "name": r[4],
            "unit": r[5],
        }

    def get_all_latest(self) -> list[dict]:
        codes = [
            r[0] for r in self.conn.execute("SELECT code FROM legacy_indicators ORDER BY code")
        ]
        return [latest for c in codes if (latest := self.get_latest(c))]

    def get_all_observations(self) -> pd.DataFrame:
        return pd.read_sql_query(
            """
            SELECT o.indicator_code, i.name, o.period, o.value,
                   o.obs_status, i.unit, i.source_agency, o.fetched_at
            FROM legacy_observations o JOIN legacy_indicators i ON o.indicator_code = i.code
            ORDER BY o.indicator_code, o.period
        """,
            self.conn,
        )

    def get_fetch_history(self, n: int = 20) -> list[dict]:
        cur = self.conn.execute(
            "SELECT indicator_code, fetched_at, rows_upserted, status, message FROM legacy_fetch_log ORDER BY id DESC LIMIT ?",
            (n,),
        )
        return [{"code": r[0], "at": r[1], "rows": r[2], "status": r[3], "msg": r[4]} for r in cur]

    def close(self):
        self.conn.close()

    def upsert_forecasts(self, rows: list[dict]) -> int:
        now = datetime.now(timezone.utc).isoformat()
        for r in rows:
            self.conn.execute(
                """
                INSERT INTO forecasts (institution, indicator, year, value, updated_at, fetched_at)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(institution, indicator, year) DO UPDATE SET
                    value=excluded.value, updated_at=excluded.updated_at,
                    fetched_at=excluded.fetched_at
            """,
                (
                    r["institution"],
                    r["indicator"],
                    r["year"],
                    r.get("value"),
                    r.get("updated_at", ""),
                    now,
                ),
            )
        self.conn.commit()
        return len(rows)

    def get_all_forecasts(self) -> pd.DataFrame:
        return pd.read_sql_query(
            """
            SELECT institution, indicator, year, value, updated_at, fetched_at
            FROM forecasts ORDER BY indicator, year, institution
        """,
            self.conn,
        )


# ─── Fetchers ─────────────────────────────────────────────────────


class NBBFetcher:
    @staticmethod
    def fetch(url: str) -> list[dict]:
        log.info(f"GET {url[:90]}...")
        resp = requests.get(url, headers=NBB_CSV_HEADER, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
        seen: dict[str, dict] = {}
        for row in csv.DictReader(io.StringIO(resp.text)):
            period = row.get("TIME_PERIOD", "").strip()
            raw = row.get("OBS_VALUE", "").strip()
            status = row.get("OBS_STATUS", "").strip()
            if not period or not raw:
                continue
            try:
                val = float(raw)
            except ValueError:
                continue
            seen[period] = {"period": period, "value": val, "obs_status": status}
        data = sorted(seen.values(), key=lambda x: x["period"])
        return data


class DBnomicsFetcher:
    @staticmethod
    def fetch(url: str, unit: str = "") -> list[dict]:
        log.info(f"GET {url[:90]}...")
        resp = requests.get(url, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()

        try:
            data = resp.json()
            series = data["series"]["docs"][0]
            periods = series["period"]
            values = series["value"]
        except (KeyError, IndexError, ValueError) as e:
            raise ValueError(f"Unexpected DBnomics JSON structure: {e}") from e

        results = []
        for p, v in zip(periods, values, strict=False):
            if str(p) < "2008":
                continue
            if v is None or v == "NA":
                continue
            try:
                val = float(v)
                results.append({"period": str(p), "value": val, "obs_status": "A"})
            except ValueError:
                continue

        if unit == "index_2010":
            results = DBnomicsFetcher._rebase_to_2010(results)
        return results

    @staticmethod
    def _rebase_to_2010(results: list[dict]) -> list[dict]:
        """Rescale values so the 2010 average equals 100 (index_2010 unit)."""
        q2010 = [r["value"] for r in results if str(r["period"]).startswith("2010")]
        if not q2010:
            return results
        avg_2010 = sum(q2010) / len(q2010)
        if avg_2010 == 0:
            return results
        return [{**r, "value": round((r["value"] / avg_2010) * 100, 2)} for r in results]


class FPBFetcher:
    INDICATORS = {1: "GDP_VOL", 3: "CPI", 5: "FISCAL_BAL"}

    @staticmethod
    def fetch(url: str = FPB_XLSX_URL) -> list[dict]:
        import tempfile

        log.info(f"GET {url[:80]}...")
        resp = requests.get(url, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(resp.content)
            tmp_path = tmp.name
        wb = load_workbook(tmp_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        year_cols = {}
        for col_offset, ind_code in FPBFetcher.INDICATORS.items():
            y1 = ws.cell(4, col_offset + 1).value
            y2 = ws.cell(4, col_offset + 2).value
            year_cols[ind_code] = [(col_offset + 1, str(int(y1))), (col_offset + 2, str(int(y2)))]
        rows = []
        for r in range(5, ws.max_row + 1):
            inst = ws.cell(r, 1).value
            if not inst or not str(inst).strip():
                continue
            upd = str(ws.cell(r, 8).value)[:10] if ws.cell(r, 8).value else ""
            for ind_code, cols in year_cols.items():
                for col_idx, year in cols:
                    val = FPBFetcher._parse_value(ws.cell(r, col_idx).value)
                    rows.append(
                        {
                            "institution": str(inst).strip(),
                            "indicator": ind_code,
                            "year": year,
                            "value": val,
                            "updated_at": upd,
                        }
                    )
        wb.close()
        Path(tmp_path).unlink(missing_ok=True)
        return rows

    @staticmethod
    def _parse_value(raw) -> float | None:
        if raw is None:
            return None
        if isinstance(raw, (int, float)):
            return round(float(raw), 2)
        s = str(raw).strip().replace(",", ".")
        if s in ("-.-", "—", "-", "...", ""):
            return None
        try:
            return round(float(s), 2)
        except ValueError:
            return None


# ─── Orchestration ────────────────────────────────────────────────


def fetch_all(db: MacroDatabase) -> bool:
    """Fetch every configured source. Returns False if any source failed."""
    all_ok = True
    for code, meta in SOURCES.items():
        db.upsert_indicator(code, meta)
        try:
            if meta.get("type") == "nbb":
                rows = NBBFetcher.fetch(meta["url"])
            else:
                rows = DBnomicsFetcher.fetch(meta["url"], meta.get("unit", ""))
            n = db.upsert_observations(code, rows)
            db.log_fetch(code, n, "OK")
            log.info(f"  OK {code}: {n} rows")
        except Exception as e:
            log.error(f"  FAIL {code}: {e}")
            db.log_fetch(code, 0, "ERROR", str(e))
            all_ok = False
    try:
        fc_rows = FPBFetcher.fetch()
        n = db.upsert_forecasts(fc_rows)
        db.log_fetch("FPB_FORECASTS", n, "OK")
    except Exception as e:
        log.error(f"  FAIL FPB_FORECASTS: {e}")
        db.log_fetch("FPB_FORECASTS", 0, "ERROR", str(e))
        all_ok = False
    return all_ok


def show_latest(db: MacroDatabase):
    latest = db.get_all_latest()
    if not latest:
        return
    print("\n" + "=" * 60)
    print("  BELGIAN MACRO DATABASE — Latest")
    print("=" * 60 + "\n")
    for e in latest:
        print(f"  {e['name']:<40} | {e['period']:<10} | {e['value']:>8.1f} {e['unit']}")


def export_data(db: MacroDatabase, fmt: str):
    df = db.get_all_observations()
    if df.empty:
        return
    out = Path(__file__).parent / "data"
    out.mkdir(parents=True, exist_ok=True)
    if fmt == "csv":
        df.to_csv(out / "belgian_macro_export.csv", index=False)
    elif fmt == "json":
        df.to_json(out / "belgian_macro_export.json", orient="records", indent=2)
    fc = db.get_all_forecasts()
    if not fc.empty:
        if fmt == "csv":
            fc.to_csv(out / "belgian_forecasts.csv", index=False)
        elif fmt == "json":
            fc.to_json(out / "belgian_forecasts.json", orient="records", indent=2)


def main():
    ap = argparse.ArgumentParser(description="Belgian Macro DB Pipeline CLI")
    ap.add_argument("--fetch", action="store_true", help="Fetch data from APIs")
    ap.add_argument("--latest", action="store_true", help="Show latest data")
    ap.add_argument("--dump", action="store_true", help="Print all data")
    ap.add_argument("--export", action="append", choices=["csv", "json"], help="Export files")
    ap.add_argument("--history", action="store_true", help="Show fetch logs")
    ap.add_argument("--db", default=str(DB_PATH), help="DB path")
    args = ap.parse_args()

    db = MacroDatabase(Path(args.db))
    if not any([args.fetch, args.latest, args.dump, args.export, args.history]):
        args.fetch = args.latest = True

    fetch_ok = True
    try:
        if args.fetch:
            log.info(f"DB: {db.db_path}")
            fetch_ok = fetch_all(db)
            if not fetch_ok:
                log.error("One or more sources failed to fetch — see fetch_log / --history")
        if args.latest:
            show_latest(db)
        if args.dump:
            df = db.get_all_observations()
            for code in df["indicator_code"].unique():
                s = df[df["indicator_code"] == code]
                print(f"\n{s.iloc[0]['name']} ({code})")
                for _, row in s.iterrows():
                    print(f"  {row['period']:<10} {row['value']:>8.1f}")
        if args.export:
            for f in args.export:
                export_data(db, f)
        if args.history:
            for e in db.get_fetch_history():
                print(f"{e['code']:<22} | {e['at'][:19]} | {e['rows']:>4} rows | {e['status']}")
    finally:
        db.close()

    if not fetch_ok:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
