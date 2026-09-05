"""
FPBSource -- fetches the FPB forecast XLSX (institution x indicator x year
scenario data, feeding the `forecasts` table -- see docs/features/
source_adapter.md, Non-goals, for why this keeps its own output shape rather
than being forced into the time-series contract).

_parse and _parse_value are FPBFetcher.fetch's original body, unchanged: the
HTTP GET, raise_for_status() and raw-bytes caching now live in the base class
(DataSource._get_with_retry / _cache_raw), which also means the original
tempfile dance is no longer needed -- the raw bytes are already in memory and
already cached to data/raw/fpb/{date}/ before _parse runs, so openpyxl reads
them directly from a BytesIO rather than a re-written temp file. The
positional column layout, year-header reading, and comma-decimal/missing-
marker value parsing are identical to before this refactor -- verified by
tests/test_fpb_source.py against a fixture workbook.
"""

import io

from openpyxl import load_workbook

from src.fetchers.base import ForecastSource

FPB_XLSX_URL = "https://www.plan.be/sites/default/files/documents/FOR_BE_FR.xlsx"


class FPBSource(ForecastSource):
    source_id = "fpb"
    adapter = "fpb"
    raw_extension = "xlsx"

    INDICATORS = {1: "GDP_VOL", 3: "CPI", 5: "FISCAL_BAL"}

    def _parse(self, raw: bytes, **kwargs) -> list[dict]:
        workbook = load_workbook(io.BytesIO(raw), data_only=True)
        sheet = workbook[workbook.sheetnames[0]]

        year_cols = {}
        for col_offset, ind_code in self.INDICATORS.items():
            y1 = sheet.cell(4, col_offset + 1).value
            y2 = sheet.cell(4, col_offset + 2).value
            year_cols[ind_code] = [
                (col_offset + 1, str(int(y1))),
                (col_offset + 2, str(int(y2))),
            ]

        rows = []
        for r in range(5, sheet.max_row + 1):
            inst = sheet.cell(r, 1).value
            if not inst or not str(inst).strip():
                continue
            updated = str(sheet.cell(r, 8).value)[:10] if sheet.cell(r, 8).value else ""
            for ind_code, cols in year_cols.items():
                for col_idx, year in cols:
                    value = self._parse_value(sheet.cell(r, col_idx).value)
                    rows.append(
                        {
                            "institution": str(inst).strip(),
                            "indicator": ind_code,
                            "year": year,
                            "value": value,
                            "updated_at": updated,
                        }
                    )
        workbook.close()
        return rows

    @staticmethod
    def _parse_value(raw) -> float | None:
        if raw is None:
            return None
        if isinstance(raw, (int, float)):
            return round(float(raw), 2)
        s = str(raw).strip().replace(",", ".")
        if s in ("-.-", "—", "-", "...", ""):
            return None
        try:
            return round(float(s), 2)
        except ValueError:
            return None
