import io

import openpyxl

from src.fetchers.fpb import FPBSource


def _build_workbook() -> bytes:
    """Mimics the real FOR_BE_FR.xlsx layout FPBSource._parse expects: year
    headers in row 4 at fixed column offsets per indicator, data from row 5,
    institution name in col 1, an 'updated' date in col 8."""
    wb = openpyxl.Workbook()
    ws = wb.active
    # Row 4: year headers for GDP_VOL (cols 2-3), CPI (cols 4-5), FISCAL_BAL (cols 6-7)
    ws.cell(4, 2, 2024)
    ws.cell(4, 3, 2025)
    ws.cell(4, 4, 2024)
    ws.cell(4, 5, 2025)
    ws.cell(4, 6, 2024)
    ws.cell(4, 7, 2025)
    # Row 5: one institution's forecasts
    ws.cell(5, 1, "FPB")
    ws.cell(5, 2, 1.5)
    ws.cell(5, 3, "1,8")
    ws.cell(5, 4, 2.1)
    ws.cell(5, 5, "-.-")
    ws.cell(5, 6, -0.5)
    ws.cell(5, 7, "n/a")
    ws.cell(5, 8, "2026-01-15")
    # Row 6: blank institution -- skipped
    ws.cell(6, 1, None)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class _FakeResponse:
    def __init__(self, content: bytes, status_code: int = 200):
        self.content = content
        self.status_code = status_code

    def raise_for_status(self):
        pass


def test_fpb_source_parses_workbook_identically_to_before_the_refactor(tmp_path, monkeypatch):
    """FPBFetcher.fetch's original behaviour, preserved verbatim: positional
    column layout, comma-decimal parsing, missing-value markers, blank
    institution rows skipped."""
    monkeypatch.setattr("src.fetchers.base.RAW_CACHE_DIR", tmp_path)
    workbook_bytes = _build_workbook()
    monkeypatch.setattr(
        "src.fetchers.base.requests.get", lambda *a, **k: _FakeResponse(workbook_bytes)
    )

    rows = FPBSource().fetch("https://example.test/fpb.xlsx", cache_key="FPB_FORECASTS")

    by_indicator_year = {(r["indicator"], r["year"]): r["value"] for r in rows}
    assert by_indicator_year == {
        ("GDP_VOL", "2024"): 1.5,
        ("GDP_VOL", "2025"): 1.8,
        ("CPI", "2024"): 2.1,
        ("CPI", "2025"): None,  # "-.-" marker
        ("FISCAL_BAL", "2024"): -0.5,
        ("FISCAL_BAL", "2025"): None,  # "n/a" unparseable
    }
    assert all(r["institution"] == "FPB" for r in rows)
    assert all(r["updated_at"] == "2026-01-15" for r in rows)
    # Only one institution row -- the blank row 6 must be skipped.
    assert len({r["institution"] for r in rows}) == 1


def test_fpb_source_caches_the_raw_workbook_bytes(tmp_path, monkeypatch):
    """The raw XLSX is now cached to disk before parsing (Block D) -- and
    that cached copy must itself be a valid workbook, not a re-encoded one."""
    monkeypatch.setattr("src.fetchers.base.RAW_CACHE_DIR", tmp_path)
    workbook_bytes = _build_workbook()
    monkeypatch.setattr(
        "src.fetchers.base.requests.get", lambda *a, **k: _FakeResponse(workbook_bytes)
    )

    FPBSource().fetch("https://example.test/fpb.xlsx", cache_key="FPB_FORECASTS")

    from datetime import date

    cached = tmp_path / "fpb" / date.today().isoformat() / "FPB_FORECASTS.xlsx"
    assert cached.read_bytes() == workbook_bytes
    # Re-parseable, confirming it's a real workbook and not corrupted by the cache write.
    reopened = openpyxl.load_workbook(cached)
    assert reopened.active.cell(5, 1).value == "FPB"
